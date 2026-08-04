import openpyxl
import json

products_wb = openpyxl.load_workbook(r'C:\Users\j.celeita\Downloads\Listado de productos _ Servicios (65).xlsx', read_only=True)
barcodes_wb = openpyxl.load_workbook(r'C:\Users\j.celeita\Belleza y Salud\COMERCIAL - Documentos\Equipo Comercial\CODIGOS DE BARRA PRODUCTOS\2026\Listado de productos- Actualizacion codigo de Barras 2026.xlsx', read_only=True)

barcode_map = {}
ws_bc = barcodes_wb['Sheet1']
for row in ws_bc.iter_rows(values_only=True, min_row=4):
    codigo = row[1]
    barcode = row[4]
    if codigo and barcode:
        barcode_map[codigo] = str(int(barcode)) if isinstance(barcode, float) else str(barcode)

ws_prod = products_wb['Sheet1']
brands = set()
products = []
for row in ws_prod.iter_rows(values_only=True, min_row=2):
    tipo, codigo, nombre, ref, categoria, distribuidor = row
    if not codigo or not nombre:
        continue
    brands.add(categoria)
    products.append({
        'codigo': codigo,
        'nombre': nombre,
        'marca': categoria,
        'codigoBarras': barcode_map.get(codigo, None),
    })

output = {'brands': sorted(list(brands)), 'products': products}
with open(r'C:\Users\j.celeita\Downloads\puntos-app\backend\products-data.json', 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print(f"Exportados {len(products)} productos de {len(brands)} marcas")
print(f"Con código de barras: {sum(1 for p in products if p['codigoBarras'])}")

products_wb.close()
barcodes_wb.close()
